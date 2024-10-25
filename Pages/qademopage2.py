import os
import sys
import logging
from openpyxl import load_workbook

# Get the directory of the current script
script_directory = os.path.dirname(os.path.abspath(__file__))

# Specify the relative path for the log directory
log_directory = "../Logs"  # Adjusted to go one level up to the "playwright" directory

# Specify the log file name
log_file_name = "qatestdemopage2.log"

# Construct the absolute path for the log file
log_file_path = os.path.join(script_directory, log_directory, log_file_name)

# Create the log directory if it doesn't exist
os.makedirs(os.path.join(script_directory, log_directory), exist_ok=True)

# Create a custom logger
logger = logging.getLogger("my_logger")
logger.setLevel(logging.INFO)  # Set the logging level to INFO

# Create a FileHandler to log to the file
log_file_handler = logging.FileHandler(log_file_path, mode="a")
log_file_handler.setFormatter(logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s"))

# Add the FileHandler to the logger
logger.addHandler(log_file_handler)

class StreamToLogger:
    def __init__(self, logger, log_level=logging.INFO):
        self.logger = logger
        self.log_level = log_level
        self.linebuf = ''

    def write(self, buf):
        for line in buf.rstrip().splitlines():
            self.logger.log(self.log_level, line)
    
    def flush(self):
        pass

# Redirect stdout and stderr to the custom logger
sys.stdout = StreamToLogger(logger, logging.INFO)
sys.stderr = StreamToLogger(logger, logging.ERROR)

class QaDemo2:
    def __init__(self, page):
        self.page = page
        #-------------------------locators--------------------
        self.elements = self.page.locator("span").filter(has_text="Elements").locator("span")
        self.textbox = page.get_by_text("Text Box")

        self.user_name = page.get_by_placeholder("Full Name")
        self.user_email = page.get_by_placeholder("name@example.com")
        self.current_address = page.get_by_placeholder("Current Address")
        self.permanant_address = self.page.locator("#permanentAddress")

        self.submit_textbox_btn = page.get_by_role("button", name="Submit")
        self.forms = self.page.get_by_text("Forms")
        self.practice_form = self.page.get_by_text("Practice Form")
        self.female_gender_radiobtn = self.page.get_by_text("Female")
        
        logger.info("QaDemo2 class initialized.")


    #-------------load url---------------------------------
    def open(self):
        logger.info("Opening the demo page.")
        self.page.goto("https://demoqa.com/sortable")


    #------------wait for the page to load-----------------
    def wait_for_page(self):
        self.page.wait_for_timeout(5000)
    

    #-----------click on Elements and textbox--------------
    def click_elements(self):
        logger.info("Clicking on the elements and text box.")
        self.elements.click(timeout=0)
        self.textbox.click(timeout=0)


    #----------fill textbox form --------------------------
    def fill_textbox_form(self, user_name, user_email, current_address, permanent_address):
        logger.info(f"Filling form with User Name: {user_name}, Email: {user_email}, "
                    f"Current Address: {current_address}, Permanent Address: {permanent_address}")
        self.user_name.click()
        self.user_name.fill(user_name)

        self.user_email.click()
        self.user_email.fill(user_email)

        self.current_address.click()
        self.current_address.fill(current_address)

        self.permanant_address.click()
        self.permanant_address.fill(permanent_address)


    #-------------submit textbox form----------------------
    def submit_textbox(self):
        logger.info("Submitting the form.")
        self.submit_textbox_btn.click(timeout=0)
        self.page.wait_for_timeout(3000)
        logger.info("Form submitted successfully.")

    #------------Add data to the fields from excel sheet----
    def fill_user_details_from_excel(self):

        logger.info(f"Loading user details from Excel file.")
        file_path = "test_data/userdata.xlsx"  # Using relative path

        file_path = os.path.join("test_data", "userdata.xlsx")
  
        wb = load_workbook(file_path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            user_name, user_email, current_address, permanent_address = row
            self.fill_textbox_form(user_name, user_email, current_address, permanent_address)
            self.submit_textbox()
            self.page.wait_for_timeout(3000)
            self.page.reload(timeout=0)
            
    #--------------closing the browser----------------------
    def close(self):
        self.page.close()
