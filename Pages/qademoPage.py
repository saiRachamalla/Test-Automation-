
from openpyxl import load_workbook
from datetime import datetime
from playwright.sync_api import expect
class QaDemo:
    def __init__(self, page):
        
        self.page = page
        self.elements=self.page.locator("span").filter(has_text="Elements").locator("span")
        self.textbox= page.get_by_text("Text Box")

        self.user_name=page.get_by_placeholder("Full Name")
        self.user_email=page.get_by_placeholder("name@example.com")

    
        self.current_address=page.get_by_placeholder("Current Address")
        self.permanant_address=page.locator("#permanentAddress")

        self.submit_textbox_btn=page.get_by_role("button", name="Submit")
        self.forms=self.page.get_by_text("Forms")
        self.practice_form=self.page.get_by_text("Practice Form")

        #practice form locators
        self.female_gender_radiobtn=self.page.get_by_text("Female")
    
    def open(self):
        self.page.goto("https://demoqa.com/sortable")
    
    def wait_for_page(self):
        self.page.wait_for_timeout(5000)

    def click_elements(self):
        self.elements.click(timeout=0)
        self.textbox.click(timeout=0)

    def fill_textbox_form(self, user_name, user_email, current_address, permanent_address):

        self.user_name.click()
        self.user_name.fill(user_name)
        self.page.wait_for_timeout(3000)

        self.user_email.click()
        self.user_email.fill(user_email)
        self.page.wait_for_timeout(3000)

        self.current_address.click()
        self.current_address.fill(current_address)
        self.page.wait_for_timeout(3000)

        self.permanant_address.click()
        self.permanant_address.fill(permanent_address)
        self.page.wait_for_timeout(3000)

    def submit_textbox(self):
        self.submit_textbox_btn.click(timeout=0)
        self.page.wait_for_timeout(3000)

    def fill_user_details_from_excel(self):
        # file_path = "C:\\Users\\poornima.r\\Documents\\Book1.xlsx"
        file_path="C:\\Users\poornima.r\\Documents\\playwright-tests\\test_data\\userdata.xlsx"
        wb = load_workbook(file_path)
        sheet = wb.active

        # Iterating over each row in the Excel sheet starting from the second row
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            user_name, user_email, current_address, permanent_address = row

            # Fill the form with the data from the Excel sheet
            self.fill_textbox_form(user_name, user_email, current_address, permanent_address)

            # Submit the form after filling details
            self.submit_textbox()

            # Adding some delay before filling the next row
            self.page.wait_for_timeout(3000)
            self.page.reload(timeout=0)
           

    def submit_textbox(self):
        self.submit_textbox_btn.click(timeout=0)
        self.page.wait_for_timeout(3000)

    def click_practice(self):
        self.forms.click(timeout=0)
        self.practice_form.click(timeout=0)

    def fill_practice_form(self):

        # Load the data from Excel
        file_path1="C:\\Users\poornima.r\\Documents\\playwright-tests\\test_data\\practice_formdata.xlsx"
        wb = load_workbook(file_path1)
        sheet = wb.active


        # Loop through each row in the sheet, starting from the second row
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):  # Skip the header row
            # Extracting values from the row
            first_name = row[0]  # First Name
            last_name = row[1]  # Last Name
            email = row[2]  # Email
            gender = row[3]  # Gender
            mobile = str(row[4])  # Mobile Number
            dob = row[5]  # Date of Birth
            subjects = row[6]
            hobbies = row[7].split(',')  # Hobbies, comma-separated
            current_address = row[9]  # Current Address
            state = row[10]  # State
            city = row[11]  # City

      
            # Parse the DOB string to extract date, month, and year
            dob_date = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
            day = dob_date.day
            month = dob_date.strftime("%B")  # Full month name
            year = dob_date.year

            
            # Filling the form
            self.page.get_by_placeholder("First Name").click()
            self.page.keyboard.type(first_name)  # Fill First Name
            self.page.get_by_placeholder("Last Name").click()
            self.page.keyboard.type(last_name)  # Fill Last Name
            self.page.get_by_placeholder("name@example.com").click()
            self.page.keyboard.type(email)  # Fill Email
            
            # Selecting gender
            if gender.lower() == "female":
                self.page.get_by_text("Female").click()
            elif gender.lower() == "male":
                self.page.get_by_text("Male", exact=True).click()
            else:
                self.page.get_by_text("Other").click()
            
            self.page.get_by_placeholder("Mobile Number").click()
            self.page.keyboard.type(mobile)  # Fill Mobile Number
    

            print(dob)
       
            self.page.locator("#dateOfBirthInput").click()
            self.page.locator("//select[@class='react-datepicker__year-select']").select_option(str(year))  # Year
            self.page.wait_for_timeout(3000)
        
            self.page.locator('//select[@class="react-datepicker__month-select"]').select_option(month)  # Month
            self.page.wait_for_timeout(3000)
            day_with_suffix = QaDemo.get_ordinal(day)

           # Convert month to its corresponding number (e.g., "January" -> 1)
            month_num = datetime.strptime(month, "%B").month

            # Get the day of the week
            day_of_week = QaDemo.get_day_of_week(year, month_num, day)

            # Use the full aria-label string
            aria_label = f"Choose {day_of_week}, {month} {day_with_suffix}, {year}"

            # Print the full aria-label to verify
            print(f'//div[@aria-label="{aria_label}"]')

            # Locate the element with the correct aria-label
            day_locator = self.page.locator(f'//div[@aria-label="{aria_label}"]')


            # Click the day
            day_locator.click()

            
            self.page.get_by_label("Select picture").set_input_files("C:\\Users\poornima.r\\Pictures\\Screenshots\\Screenshot 2024-10-23 114749.png")

            self.page.locator('//textarea[@id="currentAddress"]').click()
            self.page.keyboard.type(current_address)

            # Filling in subjects directly into the text box
            self.page.locator(".subjects-auto-complete__value-container").click()
            self.page.keyboard.type(subjects)  # Fill Subjects directly as text
            self.page.wait_for_timeout(3000)


            # Iterate through each hobby and check the corresponding checkbox
            for hobby in hobbies:
                hobby = hobby.strip().capitalize()  # Clean up the hobby name (strip spaces, capitalize first letter)
                
                print(hobby)
                if hobby=="Sports":
                    self.page.locator("//label[@for='hobbies-checkbox-1']").click()
                elif hobby=="Reading":
                    self.page.locator('//label[@for="hobbies-checkbox-2"]').click()
                else:
                    self.page.locator('//label[@for="hobbies-checkbox-3"]').click()


            
            # Selecting state
            self.page.locator("(//div[@class=' css-1wa3eu0-placeholder'])[1]").click()
            self.page.locator(f"//div[text()='{state}']").click()
            
            # # Selecting city
            # self.page.locator("#city svg").click()
            # self.page.locator(f"//div[text()='{city}']").click()
            # self.page.wait_for_timeout(3000)
            
            # Submitting the form
            self.page.get_by_role("button", name="Submit").click()

            # Optionally wait for a short period before submitting the next entry
            self.page.wait_for_timeout(2000)  # Wait for 2 seconds before filling the next form

            expect(self.page.locator('//div[@id="example-modal-sizes-title-lg"]')).to_be_visible()

            result_message=self.page.locator('//div[@id="example-modal-sizes-title-lg"]').text_content()
            
            if result_message=="Thanks for submitting the form":
                print("form data submitted successfully")

            

    def get_ordinal(day):
        if 11 <= day <= 13:  # Special case for 11th, 12th, and 13th
            return f"{day}th"
        else:
            suffixes = {1: "st", 2: "nd", 3: "rd"}
            return f"{day}{suffixes.get(day % 10, 'th')}"
        
    def get_day_of_week(year, month, day):
        date = datetime(year, month, day)
        return date.strftime("%A")  # Returns full day name like "Monday", "Tuesday"

   
    
    def close(self):
        self.page.close()
